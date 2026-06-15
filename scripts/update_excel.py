import pandas as pd
import os
import sys
import json
import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

def update_excel(file_path, data_dict):
    """
    Updates or creates an Excel file with the provided data.
    Ensures absolute data integrity and terminological accuracy.
    """
    # Updated Universal Columns
    standard_columns = [
        "Paper Title", "Year/Venue", "Supervision Mode", "Task Type",
        "Target Domain", "Model Architecture", "Knowledge Source",
        "Dataset Source", "Image-level Metrics", "Pixel-level Metrics",
        "Industrial Efficiency Metrics", "Data Strategy", "Generalization Ability",
        "Key Contributions", "Limitations & Summary"
    ]
    
    # 1. Truth-First: Ensure all fields are populated. NEVER allow blanks.
    # Populate row data, mapping both old and new names for backward compatibility during transition
    mapping = {
        "Defect & Fabric Background": "Target Domain",
        "Modality & Pre-training": "Knowledge Source",
        "Few-shot Ability": "Generalization Ability",
        "Pixel-level & Localization Metrics": "Pixel-level Metrics"
    }
    
    row_data = {}
    for col in standard_columns:
        # Get from exact name or mapped old name
        val = data_dict.get(col)
        if val is None:
            # Try to find if user sent an old-style key
            for old_key, new_key in mapping.items():
                if new_key == col:
                    val = data_dict.get(old_key)
                    break
        
        # Rigorous N/A enforcement: No blanks allowed
        if val is None or str(val).strip() == "" or str(val).lower() == "nan":
            row_data[col] = "Not Reported"
        else:
            row_data[col] = str(val).strip()

    new_row_df = pd.DataFrame([row_data])
    
    # Styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    font_content = Font(name='Arial', size=10)
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    if os.path.exists(file_path):
        try:
            wb = load_workbook(file_path)
            sheet = wb.active
            
            # Header Overwrite/Sync: Ensure the file matches the new Universal Schema
            # We rewrite Row 1 if it's an old file to ensure "Fabric" -> "Target Domain"
            for idx, col_name in enumerate(standard_columns):
                cell = sheet.cell(row=1, column=idx + 1)
                cell.value = col_name # Update to new standard headers

            next_row = sheet.max_row + 1
            
            for idx, col in enumerate(standard_columns):
                cell = sheet.cell(row=next_row, column=idx + 1)
                cell.value = row_data[col]
                cell.font = font_content
                cell.border = thin_border
                cell.alignment = alignment_center if idx < 4 else alignment_left
            
            sheet.row_dimensions[next_row].height = 40
            wb.save(file_path)
            print(f"SUCCESS: Synced schema and appended to {file_path}")
            
        except Exception as e:
            print(f"CRITICAL ERROR: {e}")
            sys.exit(1)
    else:
        try:
            # Create new with Universal Standards
            from openpyxl import Workbook
            wb = Workbook()
            sheet = wb.active
            sheet.title = "CV Anomaly Review"
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_font = Font(name='Arial', size=11, bold=True)
            
            for idx, col_name in enumerate(standard_columns):
                cell = sheet.cell(row=1, column=idx + 1)
                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment_center
                cell.border = thin_border
                sheet.column_dimensions[cell.column_letter].width = 25
            
            for idx, col_name in enumerate(standard_columns):
                cell = sheet.cell(row=2, column=idx + 1)
                cell.value = row_data[col_name]
                cell.font = font_content
                cell.border = thin_border
                cell.alignment = alignment_left
            
            wb.save(file_path)
            print(f"SUCCESS: Created new Universal database: {file_path}")
        except Exception as e:
            print(f"CRITICAL ERROR: {e}")
            sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit(1)
    target_file = sys.argv[1]
    try:
        data = json.loads(sys.argv[2])
        update_excel(target_file, data)
    except Exception as e:
        sys.exit(1)
